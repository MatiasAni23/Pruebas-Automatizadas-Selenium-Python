#importaciones necesarias para el correcto funcionamiento del código
import unittest
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from datetime import datetime, timedelta
import pygetwindow as gw
import threading
import subprocess
import time
import os




class Pruebas_retail(unittest.TestCase):
    output_folder = "" #Carpeta de salida para los videos
    CHROME_DRIVER_PATH = "" #Ruta del driver de Chrome

    CREDENTIALS = {'email': '', 'password': '', 'password2': '',  'rut': ''}

    urls = ['https://www.falabella.com/falabella-cl',
            'https://www.paris.cl',
            'https://simple.ripley.cl/']
    
    FFmpeg_COMMANDS = []

    @classmethod
    def setUpClass(cls):
        # Configuración Inicial
        chrome_options = Options()
        chrome_options.add_argument('--incognito')
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('--disable-gpu')  # Desactivar GPU
        chrome_options.add_argument('--disable-software-rasterizer')
        service = Service(executable_path="chromedriver.exe")
        
        cls.drivers = []  # Lista para los drivers
        cls.hwnds = []    # Lista para los HWND de las ventanas
        #Reportes
        
        # Archivo de log de tiempos y texto para FFmpeg
        cls.log_file = f"Reporte_Tiempos_Carga_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        cls.overlay_text_file = "overlay_text.txt"

        with open(cls.log_file, 'w') as f:
            f.write("Registro de tiempos de carga\n" + "=" * 50 + "\n")
        with open(cls.overlay_text_file, 'w') as f:
            f.write("")

        #Archivo Excel que contiene los tiempos de carga
        cls.excel_file = "Reporte_Tiempos_de_Carga.xlsx"
        cls.create_excel_report()

        
        # Carpeta para guardar los videos
        cls.test_timestamp = time.strftime('%Y%m%d_%H%M%S')
        cls.output_dir = os.path.join(cls.output_folder, f"grabacion_{cls.test_timestamp}")
        os.makedirs(cls.output_dir, exist_ok=True)  # Crear la carpeta si no existe
        
        #Archivo contenedor de acciones para los videos
        cls.ffmpeg_overlay_file = "overlay_commands.txt"

        
        # Ventana 1
        cls.driver = webdriver.Chrome(service=service, options=chrome_options)
        cls.driver.get(cls.urls[0])
        time.sleep(1)
        hwnd1 = gw.getWindowsWithTitle(cls.driver.title)[0]._hWnd
        print(f"Ventana 1 - Título: {cls.driver.title}, HWND: {hwnd1}")
        cls.drivers.append(cls.driver)
        cls.hwnds.append(hwnd1)

        # Ventana 2
        cls.driver2 = webdriver.Chrome(service=service, options=chrome_options)
        cls.driver2.get(cls.urls[1])
        time.sleep(1)
        hwnd2 = gw.getWindowsWithTitle(cls.driver2.title)[0]._hWnd
        print(f"Ventana 2 - Título: {cls.driver2.title}, HWND: {hwnd2}")
        cls.drivers.append(cls.driver2)
        cls.hwnds.append(hwnd2)

        # Ventana 3
        cls.driver3 = webdriver.Chrome(service=service, options=chrome_options)
        cls.driver3.get(cls.urls[2])
        time.sleep(1)
        hwnd3 = gw.getWindowsWithTitle(cls.driver3.title)[0]._hWnd
        print(f"Ventana 3 - Título: {cls.driver3.title}, HWND: {hwnd3}")
        cls.drivers.append(cls.driver3)
        cls.hwnds.append(hwnd3)

        # Iniciar grabación para cada ventana en un hilo
        
        cls.start_program = time.perf_counter()
        
    #Función para grabar multiples ventanas
    def grabar_multiples_ventanas(self, ventana_idx=0):
        """Inicia la grabación para una ventana específica."""
        
        windows_titles = ["Falabella", "Paris", "Ripley"]  # Títulos de las ventanas
        
        if ventana_idx < len(self.hwnds):
            hwnd = self.hwnds[ventana_idx]
            title = windows_titles[ventana_idx]  # Obtener el título correcto según el índice
            
            # Nombre del archivo de salida con el título correspondiente
            output_file = os.path.join(self.output_dir, f"grabacion_{title}_ventana_N_{ventana_idx + 1}.mp4")
            
            print(f"Iniciando grabación para ventana {ventana_idx} ({title}) con HWND={hwnd}")
            
            # Crear un hilo para iniciar la grabación
            record_thread = threading.Thread(target=self.iniciar_grabacion, args=(hwnd, output_file))
            record_thread.daemon = True
            record_thread.start()
        else:
            print(f"Índice de ventana {ventana_idx} fuera de rango.")


    #Función para iniciar grabación
    def iniciar_grabacion(self, hwnd, output_file):
        # Comandos FFmpeg usando `hwnd`
        ffmpeg_cmd = [
            "C:/webm/bin/ffmpeg", #Ruta FFmpeg (En este caso la misma del PATH)
            "-f", "gdigrab", #Formato de captura
            "-i", f"hwnd={hwnd}", #Pasar el HWND como entrada
            "-framerate", "30", #Framerate
            "-vf", (
                f"drawtext=textfile={self.overlay_text_file}:"
                f"x='main_w-text_w-10':y='main_h-text_h-50':"  # Texto en esquina inferior derecha
                f"fontsize=32:fontcolor=red:box=1:boxcolor=black@0.2:reload=1,"
                f"drawtext=fontfile=/path/to/font.ttf:"
                f"text='%{{pts\\:hms}}':"
                f"x='main_w-text_w-10':y='main_h-text_h-80':"  # Cronómetro justo encima del texto
                f"fontsize=32:fontcolor=red:box=1:boxcolor=black@0.2,"  # Cronómetro con fondo
            ),
            "-c:v", "libx264", #Codec de video
            "-preset", "ultrafast", #Velocidad de compresión
            "-crf", "28", #Calidad de compresión (Con valores más altos menor calidad de video)
            "-c:a", "aac", #Codec de audio
            "-y", #Sobreescribir el archivo de salida si ya existe
            output_file
        ]
        print(f"Iniciando grabación para hwnd={hwnd}...")
        process = subprocess.Popen(
            ffmpeg_cmd, stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE
        )
        self.FFmpeg_COMMANDS.append(process)


    #Función para detener grabación de una ventana especifica
    def detener_grabacion(self, ventana_idx):
        if ventana_idx < len(self.FFmpeg_COMMANDS):
            process = self.FFmpeg_COMMANDS[ventana_idx]
            if process:
                try:
                    process.stdin.write("q\n".encode())  
                    process.communicate(timeout=5)
                    print(f"Grabación detenida para ventana {ventana_idx}.")
                except Exception as e:
                    print(f"Error al detener FFmpeg para ventana {ventana_idx}: {e}")
                finally:
                    self.FFmpeg_COMMANDS[ventana_idx] = None

    #Función para generar reporte de tiempos en excel
    @classmethod
    def create_excel_report(cls):
        #Verificar si el archivo existe
        if os.path.exists(cls.excel_file):
            cls.workbook = openpyxl.load_workbook(cls.excel_file)
            cls.sheet = cls.workbook.active
        else: #si no existe crear uno nuevo
            cls.workbook = openpyxl.Workbook()
            cls.sheet =cls.workbook.active
            cls.sheet.title = "Reporte Tiempos de Carga"
        
        #Encabezados
        headers = ["Fecha", "Hora", "Accion", "Inicio (s)", "Fin (s)", "Duracion (s)"]
        for col, header in enumerate(headers, start = 1):
            cell = cls.sheet.cell(row = 1, column = col, value = header)
            cell.fill = PatternFill(start_color='a5f1c0', end_color = 'a5f1c0', fill_type = 'solid')
            cell.font = Font(bold = True)
            #Borde
            side = Side(border_style="thin", color="000000")
            border = Border(top=side, bottom=side, left=side, right=side)
            cell.border = border
        cls.workbook.save(cls.excel_file)
    

    def log_performance(self, action_name, start_time, end_time):
        duration = end_time - start_time
        start_seconds = start_time - self.start_program
        end_seconds = end_time - self.start_program
        current_date = datetime.now().strftime('%Y-%m-%d') #Dia
        current_time = datetime.now().strftime('%H:%M:%S') #Hora
        
        #Transformar tiempo
        start_time_transformed = f"{timedelta(seconds=int(start_seconds))}.{int((start_seconds % 1) * 100):02}"
        end_time_transformed = f"{timedelta(seconds=int(end_seconds))}.{int((end_seconds % 1) * 100):02}"

        
        #Escribir en el archivo Excel
        row = self.sheet.max_row + 1
        data = [current_date, current_time, action_name, start_time_transformed, end_time_transformed, round(duration, 2)]
        for col, value in enumerate(data, start = 1):
            self.sheet.cell(row = row, column = col, value = value) 
            
        #Guardar cambios en el archivo
        self.workbook.save(self.excel_file)
        
        #Texto que se mostrará en el reporte.txt
        message = (
            f"{action_name}\n"
            f"- Inicio accion: {start_seconds:.2f}\n"
            f"- Fin accion: {end_seconds:.2f}\n"
            f"- {action_name}: {duration:.2f} segundos \n"
        )

        # Log a archivo
        with open(self.log_file, 'a') as f:
            f.write(message + "\n")

        # **Sobrescribir el texto en overlay_text.txt**
        with open(self.overlay_text_file, 'w') as f:
            f.write(f"- {action_name}: {duration:.2f} segundos")  

        print(message)



    def test_01_carga_pagina_falabella(self):
        
        
        self.grabar_multiples_ventanas(ventana_idx=0)
        print("Test_01")
        print("Falabella")
        print("-" * 50)

        inicio_carga = time.perf_counter()
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete" #Esperar a que la página se cargue completamente
            )
            print("Pagina cargada completamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")
            
        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga

        time.sleep(2)

        #Imprimir el tiempo de carga
        print(f"El tiempo de carga de la pagina es de: {carga:.2f} segundos")
        #Almacenar el tiempo de carga en el archivo log
        self.log_performance("Carga de pagina principal - Falabella", inicio_carga, fin_carga)


    def test_02_boton_login(self):
        print("Test_02")
        print("Falabella")
        print("-" * 50)

        # Buscar el botón de inicio de sesión
        try:
            boton_login = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/header/div[2]/div/div[4]/ul/li[1]/div/div[1]/div/div[2]/p'))
            )
            print("Botón de inicio de sesión encontrado")
        except Exception as e:
            print(f"Error al buscar el botón de inicio de sesión - Falabella: {e}")
            return

        # Agregar un recuadro alrededor del botón de login
        self.driver.execute_script(
            "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
            boton_login
        )

        # Desplazar el botón de login al viewport
        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_login)

        # Mover el mouse al botón de login y mantenerlo presionado para desplegar el menú
        try:
            actions = ActionChains(self.driver)
            actions.move_to_element(boton_login).click_and_hold().perform()  # Holdear el mouse sobre el botón
            print("Mouse holdeado sobre el botón de login")
            
            # Esperar un momento para que se despliegue el menú
            time.sleep(2)

            # Buscar y hacer clic en el botón interno
            login_button = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="testId-loggedout-item-0"]'))
            )
            print("Botón interno de inicio de sesión encontrado")
            
            # Agregar un recuadro alrededor del botón interno
            self.driver.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
                login_button
            )

            # Mover el mouse al botón interno y hacer clic
            actions.move_to_element(login_button).click().perform()
            print("Clic realizado en el boton de login dentro del menú")
        except Exception as e:
            print(f"Error al buscar o hacer clic en el boton interno de inicio de sesion - Falabella: {e}")
            return
        
        
        

    def test_03_iniciar_sesion_falabella(self):
        print("Test_03")
        print("Falabella")
        print("-" * 50)

        # Ingresar email
        try:
            ingresar_email = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="testId-cc-login-form-email-input"]'))
            )
            
            # Resaltar el campo
            self.driver.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
                ingresar_email
            )
            # Ingresar el email
            ingresar_email.send_keys(self.CREDENTIALS['email'])
            print("El email se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar el email: {e}")
            return

        # Ingresar contraseña
        try:
            ingresar_contrasena = WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="testId-cc-login-form-password-input"]'))
            )
            
            # Resaltar el campo
            self.driver.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
                ingresar_contrasena
            )
            # Ingresar la contraseña
            ingresar_contrasena.send_keys(self.CREDENTIALS['password2'])
            print("La contraseña se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar la contraseña: {e}")
            return

        # Presionar botón Iniciar sesión
        try:
            boton_iniciar_sesion = WebDriverWait(self.driver, 10).until(
                EC.element_to_be_clickable((By.ID, 'testId-cc-login-form-submit'))
            )
            
            # Resaltar el botón
            self.driver.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
                boton_iniciar_sesion
            )
            # Hacer scroll al botón
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_iniciar_sesion)
            
            # Clic en el botón
            inicio_carga = time.perf_counter()
            boton_iniciar_sesion.click()
            print("El botón de iniciar sesión se ha presionado correctamente")
        except Exception as e:
            print(f"Error al presionar el botón de iniciar sesión: {e}")
            return
        #Esperar a que cambie la url
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: d.current_url != self.urls[0]
            )
            print("Redirección detectada")
        except Exception as e:
            print(f"Error al detectar la redirección: {e}")

        #Medir tiempo de carga de la página
        try:
            WebDriverWait(self.driver, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente")

        except Exception as e:
            print(f"Error al cargar la pagina: {e}")
        
        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga
        print(f"El tiempo de carga de la pagina es de: {carga:.2f} segundos")

        #Almacenar el tiempo de carga en el archivo log
        self.log_performance("Carga de pagina luego de presionar el boton de inicio de sesion - Falabella", inicio_carga, fin_carga)

        time.sleep(2)

    
    def test_04_busqueda_producto(self):
        print("Test_04")
        print("Falabella")
        print("-" * 50)

        #Buscar producto
        try:
            buscar_producto = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="testId-SearchBar-Input"]'))
            )
            buscar_producto.send_keys("Playstation 5")
            buscar_producto.send_keys(Keys.ENTER)
            print("Producto encontrado")
        except Exception as e:
            print(f"Error al buscar el producto: {e}")

        
        
        self.detener_grabacion(ventana_idx=0)
        
        time.sleep(3)
        
        


    def test_05_carga_pagina_paris(self):
        
        self.grabar_multiples_ventanas(ventana_idx=1)
        print("Test_01")
        print("Paris")
        print("-" * 50)

        inicio_carga = time.perf_counter()
        try:
            WebDriverWait(self.driver2, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete" #Esperar a que la página se cargue completamente
            )
            print("Página cargada completamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")
            
        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga

        #Imprimir el tiempo de carga
        print(f"El tiempo de carga de la pagina de Paris.cl es de: {carga:.2f} segundos")
        #Almacenar el tiempo de carga en el archivo log
        self.log_performance("Carga de pagina principal - Paris.cl", inicio_carga, fin_carga)

        time.sleep(3)



    def test_06_boton_login_paris(self):
        print("Test_02")
        print("Paris")
        print("-" * 50)

        #Buscar el botón de inicio de sesión
        try:
            boton_login = WebDriverWait(self.driver2, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/header/div[2]/div/nav/div[4]/div/div[1]/div/button/div[2]/div[2]'))
            )
        except Exception as e:
            print(f"Error al buscar el botón de inicio de sesion - Paris.cl: {e}")


        self.driver2.execute_script(
            "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
            boton_login
        )
        
        self.driver2.execute_script("arguments[0].scrollIntoView({block: 'center'});", boton_login)
        
        # Esperar un momento para que se despliegue el menú
        time.sleep(1)

        
        try:
            
            actions = ActionChains(self.driver2)
            actions.move_to_element(boton_login).click_and_hold().perform()
            print("Mouse holdeado sobre el botón de inicio de sesión")
            
            
            time.sleep(2)
            
            login_button = WebDriverWait(self.driver2, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/header/div[2]/div/nav/div[4]/div/div[1]/div/div/div/div/div/div[1]/button'))
            )
            print("Botón de inicio de sesión encontrado")
            
            self.driver2.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                login_button
            )
            
            actions.move_to_element(login_button).click().perform()
            print("Clic realizado en el botón de inicio de sesión")
        except Exception as e:
            print(f"Error al buscar el botón de inicio de sesion - Paris.cl: {e}")
            return

        time.sleep(3)


    def test_07_iniciar_sesion_paris(self):

        print("Test_03")
        print("Paris")
        print("-" * 50)

        #Ingresar email
        try:
            ingresar_email = WebDriverWait(self.driver2, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="dwfrm_login_username"]'))
            )
            #Resaltar el campo
            self.driver2.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                ingresar_email
            )
            ingresar_email.send_keys(self.CREDENTIALS['email'])
            print("El email se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar el email: {e}")

        #Ingresar contraseña
        try:
            ingresar_contrasena = WebDriverWait(self.driver2, 10).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="dwfrm_login_password"]'))
            )
            
            self.driver2.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                ingresar_contrasena
            )
            
            ingresar_contrasena.send_keys(self.CREDENTIALS['password'])
            print("La contraseña se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar la contraseña: {e}")

        #Presionar botón Iniciar sesión
        try:
            boton_iniciar_sesion = WebDriverWait(self.driver2, 10).until(
                EC.element_to_be_clickable((By.XPATH, '/html/body/div[17]/div[2]/div/form/div[4]'))
            )
            inicio_carga = time.perf_counter()
            
            self.driver2.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                boton_iniciar_sesion
            )
            
            boton_iniciar_sesion.click()
            
            print("El botón de iniciar sesion se ha presionado correctamente")
        except Exception as e:
            print(f"Error al presionar el boton de iniciar sesion: {e}")

        #Esperar a que cambie la url
        try:
            WebDriverWait(self.driver2, 10).until(
                lambda d: d.current_url != self.urls[1]
            )
            print("Redirección detectada")
        except Exception as e:
            print(f"Error al detectar la redirección: {e}")

        #Medir tiempo de carga de la página
        try:
            WebDriverWait(self.driver2, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")

        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga
        print(f"El tiempo de carga de la pagina es de: {carga:.2f} segundos")

        #Almacenar el tiempo de carga en el archivo log
        self.log_performance("Carga de pagina luego de presionar el boton de inicio de sesion - Paris.cl", inicio_carga, fin_carga)

        time.sleep(3)
        

    def test_08_busqueda_producto(self):
        print("Test_04")
        print("Paris")
        print("-" * 50)

        #Buscar producto

        try:
            buscar_producto = WebDriverWait(self.driver2, 10).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="desktop-search__form-input"]'))
            )
            
            self.driver2.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                buscar_producto
            )
            
            buscar_producto.send_keys("Audifonos Bluetooth")
            buscar_producto.send_keys(Keys.ENTER)
            print("Producto encontrado")
        except Exception as e:
            print(f"Error al buscar el producto: {e}")


        
        self.detener_grabacion(ventana_idx=1)
        
        time.sleep(3)

    def test_09_carga_pagina_ripley(self):
        
        self.grabar_multiples_ventanas(ventana_idx=2)
        print("Test_01")
        print("Ripley")
        print("-" * 50)

        inicio_carga = time.perf_counter()
        try:
            WebDriverWait(self.driver3, 10).until(
                lambda d: d.execute_script("return document.readyState") == "complete" #Esperar a que la página se cargue completamente
            )
            print("Página cargada completamente")
        except Exception as e:
            print(f"Error al cargar la pagina - Ripley.cl: {e}")
            
        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga

        #Imprimir el tiempo de carga
        print(f"El tiempo de carga de la pagina es de: {carga:.2f} segundos")
        #Almacenar el tiempo de carga en el archivo log
        self.log_performance("Carga de pagina principal - Ripley.cl", inicio_carga, fin_carga)


        time.sleep(1)

    def test_10_iniciar_sesion(self):
        
        
        print("Test_02")
        print("*" * 50)
        inicio_boton = time.perf_counter()
        #presionar botón iniciar sesión
        
        try:
            presionar_boton = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/main/section/nav/main/nav/section[2]/div[1]/section/button"))
            )
            
            # Resaltar el campo
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';", 
                presionar_boton
            )
            
            presionar_boton.click()
            print("El botón se ha presionado correctamente")
        except Exception as e:
            print(f"Error al presionar el botón: {e}")



        fin_boton = time.perf_counter()
        carga_boton = fin_boton - inicio_boton
        print(f"El tiempo de carga luego de presionar el boton de inicio de sesion hasta el formulario es de: {carga_boton:.2f} segundos")

        #Almacenar resultados
        self.log_performance("Carga de pagina luego de presionar el boton de inicio de sesion hasta el formulario - Ripley.cl", inicio_boton, fin_boton)


        time.sleep(1)


        #Ingresar rut
        try:
            ingresar_rut = WebDriverWait(self.driver3, 30).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="identifier"]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                ingresar_rut
            )
            
            ingresar_rut.send_keys(self.CREDENTIALS['rut']) 
            print("El rut se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar el rut: {e}")


        #Ingresar contraseña
        try:
            ingresar_contrasena = WebDriverWait(self.driver3, 30).until(
                EC.visibility_of_element_located((By.XPATH, '//*[@id="password"]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                ingresar_contrasena
            )
            
            ingresar_contrasena.send_keys(self.CREDENTIALS['password'])
            print("La contraseña se ha ingresado correctamente")

        except Exception as e:
            print(f"Error al ingresar la contraseña: {e}")

        #Presionar botón Iniciar sesión

        try:
            boton_iniciar_sesión = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login"]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                boton_iniciar_sesión
            )
            
            
            inicio_carga = time.perf_counter()  
            boton_iniciar_sesión.click()
            print(f"El boton de iniciar sesion se ha presionado correctamente")
            

        except Exception as e:
            print(f"Error al presionar el boton de iniciar sesion: {e}")
            
        

        #Esperar a que cambie la url

        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.current_url != self.urls[2]
            )
            print("Redirección detectada")
        except Exception as e:
            print(f"Error al detectar la redirección: {e}")
            
        
        #Medir tiempo de carga de la pagina

        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente - Ripley.cl")
        except Exception as e:
            print(f"Error al cargar la pagina: {e}")
            

        fin_carga = time.perf_counter()
        carga = fin_carga - inicio_carga
        print(f"El tiempo de carga de la pagina luego de presionar el boton de inicio de sesion es de: {carga:.2f} segundos")

        #Almacenar resultados
        self.log_performance("Carga de pagina luego de presionar el boton de inicio de sesion", inicio_carga, fin_carga)

        time.sleep(3)

    def test_11_buscar_producto(self):

        
        print("Test_03")
        print("*" * 50)
        #Presionar barra de busqueda
        try:
            presionar_barra = WebDriverWait(self.driver3,  30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/main/section/nav/main/nav/div/form/input'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_barra
            )
            
            presionar_barra.click()
            print("Se ha presionado correctamente la barra de búsqueda")
            time.sleep(2)
        except Exception as e:
            print(f"Error al presionar la barra de búsqueda: {e}")

        #Ingresar datos o producto a buscar

        try:
            buscar_producto = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/main/section/nav/main/nav/div/form/input'))
            )
            buscar_producto.send_keys("Layton Parfums de Marly")
            time.sleep(2)
            inicio_busqueda = time.perf_counter()
            buscar_producto.send_keys(Keys.ENTER)
            print("El producto a buscar se ha ingresado correctamente")
        except Exception as e:
            print(f"Error al ingresar el producto a buscar: {e}")

        fin_busqueda = time.perf_counter()
        carga_busqueda = fin_busqueda - inicio_busqueda
        print(f"La carga de la búsqueda es de: {carga_busqueda:.2f} segundos")

        #Almacenar resultados
        self.log_performance("Carga de pagina al realizar busqueda", inicio_busqueda, fin_busqueda)


        time.sleep(2)

#Testeo para seleccionar producto y cuanto se demora en transición de paginas

    def test_12_seleccionar_producto(self):

        print("Test_04")
        print("*" * 50)


        try:
            seleccionar_producto = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[9]/div[1]/div/div[1]/div[3]/section/div/div/div[1]/div/a/div[3]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                seleccionar_producto
            )
            
            inicio_seleccion = time.perf_counter()
            seleccionar_producto.click()
            print("Se seleccionó el producto correctamente")
        except Exception as e:
            print(f"No se pudo seleccionar el producto: {e}")


        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La pagina se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")
            

        fin_seleccion = time.perf_counter()
        carga = fin_seleccion - inicio_seleccion
        print(f"El tiempo de carga de la pagina luego de seleccionar el producto y cargar la pagina: {carga:.2f} segundos")

        #Almacenar resultados
        self.log_performance("Carga de pagina luego de presionar el producto y cargar la pagina", inicio_seleccion, fin_seleccion)

        time.sleep(2)

    #Testep para agregar productos al carrito

    def test_13_agregar_carrito(self):

        print("Test_05")
        print("*" * 50)

        try:

            agregar_producto = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="buy-button"]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                agregar_producto
            )
            
            agregar_producto.click()
            print("Se agregó el producto al carrito")
            time.sleep(3)

        except Exception as e:
            print(f"No se pudo agregar el producto al carrito: {e}")
    

    # Testeo para ver el perfil y cuanto se demora cada componente dentro de la pagina

    def test_14_ver_perfil(self):

        print("Test_06")
        print("*" * 50)

        time.sleep(5)

        #Click en botón de cuenta
        try:
            presionar_boton = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[5]/header/section/nav/div/div[2]/div[1]/div/div/a/div'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_boton
            )
            
            presionar_boton.click()
            print("El boton de cuenta se ha presionado correctamente")
        except Exception as e:
            print(f"Error al presionar el boton: {e}")


        #Click para redirigir a la cuenta

        try:
            presionar_cuenta = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, "//li[@class='credentials__my-account-link']/a[text()='Mi cuenta']"))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_cuenta
            )
            inicio_cuenta = time.perf_counter()
            presionar_cuenta.click()
            print("Se ha presionado la cuenta correctamente")
        except Exception as e:
            print(f"Error al presionar la cuenta: {e}")
        

        #Carga componente Mis datos personales
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[2]/div[2]/div[2]/div"))
            )
            print("El componente 'Mis datos personales' se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar el componente 'Mis datos personales': {e}")


        fin_datos = time.perf_counter()
        carga_datos = fin_datos - inicio_cuenta


        #Carga componente contenedor izquierdo
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div'))
            )
            print("El contenedor izquierdo se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar el contenedor: {e}")

        fin_contenedor = time.perf_counter()
        carga_contenedor = fin_contenedor - inicio_cuenta


        #Carga ripley puntos
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[1]/a/div'))
            )
        except Exception as e:
            print(f"Error al cargar el contenedor de ripley puntos: {e}")

        fin_puntos = time.perf_counter()
        carga_puntos = fin_puntos - inicio_cuenta


        #Carga página completa
        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La pagina se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la pagina: {e}")
        

        fin_cuenta = time.perf_counter()
        tiempo_carga = fin_cuenta - inicio_cuenta
        

        #Imprimir resultados
        print(f"El tiempo de carga del componente 'Mis datos personales' es de: {carga_datos:.2f} segundos")
        print(f"El tiempo de carga del contenedor izquierdo es de: {carga_contenedor:.2f} segundos")
        print(f"El tiempo de carga del contenedor de ripley puntos es de: {carga_puntos:.2f} segundos")
        print(f"El tiempo de carga despues de presionar el boton de cuenta es de: {tiempo_carga:.2f} segundos")

        #Almacenar resultados
        self.log_performance("El tiempo de carga del componente 'Mis datos personales'", inicio_cuenta, fin_datos)
        self.log_performance("El tiempo de carga del contenedor izquierdo", inicio_cuenta, fin_contenedor)
        self.log_performance("El tiempo de carga del contenedor de ripley puntos", inicio_cuenta, fin_puntos)
        self.log_performance("Carga de pagina luego de presionar el boton de cuenta", inicio_cuenta, fin_cuenta)

        time.sleep(2)


    # Testeos de los tiempos de carga para los elementos de la barra lateral

    def test_15_click_elemento_compras(self):

        print("Test_07")
        print("*" * 50)
        
        #Click en compras realizadas
        try:
            presionar_compras = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[1]/a/span[1]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_compras
            )
            
            inicio_compras = time.perf_counter()
            presionar_compras.click()
            print("Se ha presionado compras realizadas correctamente")
        except Exception as e:
            print(f"Error al presionar compras realizadas: {e}")



        #Carga contenedores de productos comprados
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div/div[2]/div[1]/div'))
            )
        except Exception as e:
            print(f"Error al cargar el contenedor de productos comprados : {e}")

        fin_compras = time.perf_counter()
        carga_compras = fin_compras - inicio_compras

        print(f"El tiempo de carga de productos comprados es de: {carga_compras:.2f} segundos")

        

        #Carga de elementos dentro del contenedor de productos comprados
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div/div[2]/div[1]/div/div[2]/div/div/div/div/div/div[3]')) #contenedor de datos (fecha retiro, direccion y quien retira)
            )

        except Exception as e:
            print(f"Error al cargar los elementos dentro del contenedor de productos comprados : {e}")

        fin_elementos = time.perf_counter()
        carga_elementos = fin_elementos - inicio_compras

        print(f"El tiempo de carga de los elementos dentro del contenedor de productos comprados es de: {carga_elementos:.2f} segundos")

        


        #Carga página completa
        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")

        fin_carga_seccion_compras = time.perf_counter()
        carga_seccion_compras = fin_carga_seccion_compras - inicio_compras

        print(f"El tiempo de carga de la seccion de compras es de: {carga_seccion_compras:.2f} segundos")



        #Almacenar resultados
        self.log_performance("Carga de productos comprados", inicio_compras, fin_compras)
        self.log_performance("Carga de elementos dentro del contenedor de productos comprados", inicio_compras, fin_elementos)
        self.log_performance("Carga de la seccion de compras", inicio_compras, fin_carga_seccion_compras)

        time.sleep(2)

    def test_16_click_solicitudes(self):


        print("Test_08")
        print("*" * 50)

        #click en solicitudes de atención
        try:
            presionar_solicitudes = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[2]/a[1]/span[1]'))
            )
            
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_solicitudes
            )
            inicio_solicitudes = time.perf_counter()
            presionar_solicitudes.click()
            print("Se ha presionado solicitudes de atencion correctamente")
        except Exception as e:
            print(f"Error al presionar solicitudes de atencion: {e}")


        #Carga contenedores de solicitudes de atención

        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div/div/div[3]/div[2]/div'))
            )
        except Exception as e:
            print(f"Error al cargar el contenedor de solicitudes de atención : {e}") 

        fin_carga_cont_solicitudes = time.perf_counter()
        carga_cont_solicitudes = fin_carga_cont_solicitudes - inicio_solicitudes

        print(f"El tiempo de carga de solicitudes de atencion es de: {carga_cont_solicitudes:.2f} segundos")


        #Carga completa de la pagina
        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")

        fin_solicitudes = time.perf_counter()
        carga_solicitudes = fin_solicitudes - inicio_solicitudes

        print(f"El tiempo de carga de la pagina luego de presionar solicitudes de atencion es de: {carga_solicitudes:.2f} segundos")


        #Almacenar Resultados
        self.log_performance("Carga de solicitudes de atencion", inicio_solicitudes, fin_carga_cont_solicitudes)
        self.log_performance("Carga de la pagina luego de presionar solicitudes de atencion", inicio_solicitudes, fin_solicitudes)


        time.sleep(2)
            


    def test_17_click_ripley_puntos(self):
        
        print("Test_09")
        print("*" * 50)



        #Click en ripley puntos Go

        try:
            presionar_ripley_puntos = WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[1]/div/div[2]/div[4]/a[1]/span[1]'))
            )
            
            self.driver3.execute_script(
                "arguments[0].style.border='3px solid red'; arguments[0].style.transition='border 0.3s ease-in-out';",
                presionar_ripley_puntos
            )
            
            inicio_ripley_puntos = time.perf_counter()
            presionar_ripley_puntos.click()
            print("Se ha presionado ripley puntos correctamente")

        except Exception as e:
            print(f"Error al presionar ripley puntos: {e}")


        #Carga contenedores de ripley puntos

        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[1]/div[2]'))
            )
        except Exception as e:
            print(f"Error al cargar el contenedor de ripley puntos : {e}")

        fin_carga_cont_ripley_puntos = time.perf_counter()
        carga_cont_ripley_puntos = fin_carga_cont_ripley_puntos - inicio_ripley_puntos

        print(f"El tiempo de carga de ripley puntos es de: {carga_cont_ripley_puntos:.2f} segundos")


        #Carga de canjeo ripley puntos
        try:
            WebDriverWait(self.driver3, 30).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div[2]/div[2]/div[2]'))
            )
        except Exception as e:
            print(f"Error al cargar el contenedor de ripley puntos : {e}")

        fin_canjeo_ripley_puntos = time.perf_counter()
        carga_cont_ripley_puntos = fin_canjeo_ripley_puntos - inicio_ripley_puntos

        print(f"El tiempo de carga de ripley puntos es de: {carga_cont_ripley_puntos:.2f} segundos")




        #Carga completa de la pagina
        try:
            WebDriverWait(self.driver3, 30).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            print("La página se ha cargado correctamente")
        except Exception as e:
            print(f"Error al cargar la página: {e}")

        fin_ripley_puntos = time.perf_counter()
        carga_ripley_puntos = fin_ripley_puntos - inicio_ripley_puntos

        print(f"El tiempo de carga de la pagina luego de presionar ripley puntos es de: {carga_ripley_puntos:.2f} segundos")

        #Almacenar resultados
        self.log_performance("Carga de ripley puntos", inicio_ripley_puntos, fin_carga_cont_ripley_puntos)
        self.log_performance("Carga de contenedor de ripley puntos", inicio_ripley_puntos, fin_canjeo_ripley_puntos)
        self.log_performance("Carga de la pagina luego de presionar ripley puntos Go", inicio_ripley_puntos, fin_ripley_puntos)


        time.sleep(2)
        self.detener_grabacion(ventana_idx=2)



    @classmethod
    def tearDownClass(cls):
        for driver in cls.drivers:
            driver.quit()
        
        print("Todos los navegadores se han cerrado correctamente.")


if __name__ == '__main__':
    unittest.main()

        
        
